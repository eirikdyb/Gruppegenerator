# -*- coding: utf-8 -*-

from pylab import randint
from openpyxl import load_workbook
import sys

def rand_grp(names, grp_size):
    #Funksjon som tar inn en liste med navn og en int som beskriver gruppestørrelse,
    #og lager tilfeldige grupper av den størrelsen. Hvis det ikke går opp
    #blir de resterende plassert i grupper (starter på gruppe 1).

    grp_list = []
    num_grps = len(names) // grp_size
    rest_ppl = len(names) % grp_size
    
    while len(names) >= grp_size:       #Lager nye grupper med grp_size tilfeldige elementer fra names
        new_grp = []
        for i in range(grp_size):
            new_grp.append(names.pop(randint(len(names))))
        grp_list.append(new_grp)
        
    
    for i in range(rest_ppl):           #Fyller inn de resterende personene
        grp_list[i % num_grps].append(names.pop(0))

    return grp_list
    
def print_grp(grp_list):
    #Skrive ut gruppelistene
    i = 1
    for grp in grp_list:
        print(f"Gruppe {i}: ", end ="")
        for name in grp:
            if name == grp[-2]:
                print(f"{name} ", end ="")
            elif name == grp[-1]:
                print(f"og {name}.")
            else:
                print(f"{name}, ", end ="")
        i += 1
        
def print_grp_txt(grp_list):
    f = open("Grupper.txt", "w")
    i = 1
    for grp in grp_list:
        f.write(f"Gruppe {i}: ")
        for name in grp:
            if name == grp[-2]:
                f.write(f"{name} ")
            elif name == grp[-1]:
                f.write(f"og {name}.\n")
            else:
                f.write(f"{name}, ")
        i += 1
    f.close()
        
def read_name_txt(filename): #Henter inn første element fra hver rad i en navneliste.
    names = []
    with open(filename,'r') as namelist:
        for line in namelist:
            names.append(line.split(None, 1)[0])

    return names

def read_name_xls(filename): #Henter inn første kolonne i excel-arket. Tittel i første rad.
    names = []    
    workbook = load_workbook(filename)
    ws = workbook[workbook.sheetnames[0]]

    for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row, values_only=True):
        for value in row:
            if value:
                names.append(value)
            
    return names
    
def main():
    print("Skriv inn filnavn. Husk .xls eller .xlsx")
    filnavn = str(input())
    
    if len(sys.argv) == 1:
        grp_size = int(input("Skriv inn gruppestørrelse: "))
    elif len(sys.argv) == 2:
        grp_size = int(sys.argv[-1])
    else:
        print("Gruppestørrelse ikke definert")
        print("Bruker standardstørrelse 3")
        grp_size = 3
        filnavn = "elever.xlsx"
    
    names = read_name_xls(filnavn)
    grps = rand_grp(names,grp_size)
    
    print_to_console = True
    print_to_txt = True
    
    if print_to_txt:
        print_grp_txt(grps)
    if print_to_console:
        print_grp(grps)

main()